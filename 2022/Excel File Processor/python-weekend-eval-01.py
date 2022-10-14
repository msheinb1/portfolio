from concurrent.futures import process
from openpyxl import load_workbook
from datetime import datetime, date
import os, re, logging

excluded_dirs = ['archived_files', 'error_files']
months = ['january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december']
error_files = []
processed_files = None
abnormal_exit = 'Program exitted abnormally.'

#The program only searches 1 directory.
#Assumes any .xlsx file in the selected is a file to be processed, and ONLY .xlsx files
#Will place logfile in the current directory.
def find_all(pattern, path):
    global error_files
    file_list = []
    for root, dirs, files in os.walk(path, topdown=True):
        for name in files:
            if re.search(pattern, name):
                file_list.append(os.path.join(root, name))
            else:
                error_files.append(os.path.join(root, name))
    return file_list

def contains_month(my_string):
    global error_files
    
    index = -1
    for x in months:
        index = os.path.basename(my_string).find(x)
        if (index > -1):
            return True
    error_files.append(my_string)    
    return False

#returns tuple of (int, string): (index, month_name)
def get_month(my_string):
    index = -1
    for idx, x in enumerate(months):
        index = os.path.basename(my_string).lower().find(x[:3])
        if(index > -1):
            return (idx, x[:3])
        else:
            continue
    return None

def move_file(file, folder_name, cwd, file_descriptor):
    try:
        os.rename(file, cwd +  '\\' + folder_name + '\\' + os.path.basename(file))
        logging.info('Moved {} file to {}: {}'.format(file_descriptor, folder_name, file))
    except:
        logging.error('Could not move {} file to {}: {}'.format(file_descriptor, folder_name, file))

def good_bad(num, threshold):
    if num > threshold:
        return 'good'
    elif num < threshold:
        return 'bad'
	return 'neutral'

if __name__ == "__main__":
    #Tries to make logfile if does not exist.
    #If the logfile exists it will just append to existing logfile.
    assert(type(float) == type)
    exit(0)
    today = date.today()
    logfile = '{}-logfile.log'.format(today)
	if(os.path.exists(logfile) == True):
        open(logfile, 'x')
    else:
        print('Logfile already exists. Appending to existing log.')
    
    try:
        logging.basicConfig(filename= logfile, encoding='utf-8', level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
        logging.info('Program Started')
    except:
        print('Error: Could not open log file. Exiting')
        exit(1)

    cwd = None
    user_input = input('Enter target directory (Use * for current directory. Do not include trailing backslash):\n')

    if user_input == '*':
        try:
            cwd = os.getcwd()
        except:
            logging.error('Could not find the directory {}'.format(cwd))
            logging.error(abnormal_exit)
            print('Error: Could not find the directory {} {}\nTerminating program.'.format(cwd))
            exit(1)
    else:
        cwd = user_input
        if(os.path.exists(cwd) == False):
            logging.error('Could not find the directory {}'.format(cwd))
            logging.error(abnormal_exit)
            print('Error: Could not find the directory {}\nTerminating program.'.format(cwd))
            exit(1)
    
    logging.info('Directory to search is set to: ' + cwd)

    print('Enter files to process (Use * for all in listed directory,):\n- Use q to stop entering files')

    file_list = []
    user_input = ''
    while user_input != 'q':
        user_input = input('')
        if user_input == '*':
            regex_pattern = '^.*_[0-9]+.xlsx'
            
            #Find all files of the naming format in the current directory: expedia_report_monthly_*_*.xlsx
            try:
                file_list = find_all(regex_pattern, cwd)
            except:
                logging.error('Could not open the given directory: {}'.format(cwd))
                logging.error(abnormal_exit)
                print('Error: Could not open the given directory: {}\nTeriminating program.'.format(cwd))
                exit(1)
            break
        elif user_input == 'q':
            break
        else:
            file_list.insert(len(file_list), cwd + '\\' + user_input)
    
    logging.info('Added {} files to processing queue.'.format(len(file_list)))
    logging.info('Added {} files to error queue.'.format(len(error_files)))
            
    #Filters all files that do not have a month as the first asterisk
    #It ONLY checks to see if a month is listed, the rest of the format should be covered by the previous search
    #I iterated over the file_list to check for the month separately because it sounds like it would have gotten VERY complicated in regex, if it was even possible in regex
    prev_len = len(error_files)
    months_iterator = filter(contains_month, file_list)
    file_list = list(months_iterator)
    num_changed = len(error_files) - prev_len
    if num_changed > 0:
        logging.info('Moved {} files from processing queue to error queue.'.format(num_changed))
    
    #The error_file list was then iterated over to remove non .xlsx documents
    prev_len = len(error_files)
    error_files = list(filter(lambda x: re.search('.*.xlsx$', x), error_files))
    num_changed = prev_len - len(error_files)
    if num_changed > 0:
        logging.info('Removed {} non .xlsx files from error queue.'.format(num_changed))

    #Removes files from the archived_files + error_files directory
    #Will erroneously not process any files with the correct file format that are not in the processed file list if they are placed in a directory named archived_files or error_files
    prev_len = len(file_list)
    prev_len_2 = len(error_files)
    for x in excluded_dirs:
        file_list = [s for s in file_list if x not in s] 
        error_files = [s for s in error_files if x not in s]
    num_changed = prev_len - len(file_list)
    if num_changed > 0:
        logging.info('Removed {} files from processing queue that are in the archived_files directory.'.format(num_changed))
    num_changed = prev_len_2 - len(error_files)
    if num_changed > 0:
        logging.info('Removed {} files from processing queue that are in the error_files directory.'.format(num_changed))

    #Get files that are already processed from file.lst
    #Assumes file.lst is in the current working directory, and exists
    #File.lst does not use the absolute paths of the files since the program itself moves processed files to a separate folder
    #It puts the processed files into a set first in case there are duplicates in file.lst due to some kind of error.
    #It then filters out any non .xlsx files in the processed list
    #Will maybe make the program create a blank file.lst if I have time.

	if(os.path.exists('file.lst') == True):
        processed_file_log = open('file.lst', 'r')
    else:
        try:
            open('file.lst', 'x')
            logging.info('Created empty file: file.lst')
            processed_file_log = open('file.lst', 'r')
        except:
            logging.error('Could not create file: file.lst')
            logging.error(abnormal_exit)
            print('Could not open or create list of already processed files.\nTeriminating program.')
            exit(1)
    processed_files = set(processed_file_log.read().split('\n'))
    processed_files = set(filter(lambda x: re.search('.*.xlsx$', x), processed_files))
    processed_file_log.close()

    #Compare file_list to results, filtering out ones that have been processed.
    #They get removed from file_list completely because they have been properly processed.
    #They are not sent to error_files!
    #If there is a match the corresponding file it attempts to move the file to archived files
    #If the file cannot be moved (such as due to a file with the same name existing in the destination),
    #It leaves the file where it is and does not rewrite the file.
    #This way, either user or another program can make sure that the files are actually the same and not different files accidentally named erroneously.
    temp = None
    for x in processed_files:
        temp = [s for s in file_list if x in s]
        if len(temp) > 0:
            move_file(temp[0], 'archived_files', cwd, 'already processed')
            file_list.remove(temp[0])

    #Moving Error Files
    #Checking to see file exists first
    #Decided to move error_files to their folder here so theyh could be logged properly.
    #For example: a user gives an incorrectly named .xlsx file that does not actually exist in system.
    file_name = None
    for x in error_files:
        if(os.path.exists(x) == False):
            logging.error('File does not exist: {}'.format(x))
        else:
            move_file(x, 'error_files', cwd, 'incorrectly named')

    #Process file_list.
    #I decided to move each item in file_list as I processed the item so the correct log message can be added
    #Check to ensure correct number of sheets.
    #Assumes data is in the correct cells.
	#Can and will break spectacularly if the sheet template is not exactly as given.
    #Since the Excel sheet is supposed to be 12 month rolling format, only checks 12 rows or columns back.
    #Throws error if can't match it.

	#Other things to consider when processing files:
	#Grab the data associated with the first instance of the month, even if that month appears twice.
	# - Above behavior can possibly raise an error if month is incorrectly applied early, maybe check to see if that can be handled later if I have time.
	#Log error if the month doesn't appear, move file to error folder
    #I don't think I have time to fix it.

    num_processed = 0
    if len(file_list) <= 0:
        logging.info('No files available to process.')
        logging.info('Program exited normally.')
        exit(0)

    processed_file_log = open('file.lst', 'a')
    for x in file_list:
        current_file_log = None
        current_workbook = None
        active_sheet = None
        selected_month = None
        has_error = False
        if(os.path.exists(x) == False):
            logging.error('File does not exist: {}'.format(x))
        else:
            try:
                current_workbook = load_workbook(filename = x)
            except:
                logging.error('Could not open workbook: {}'. format(x))
                move_file(x, 'error_files', cwd, 'unopenable')
                continue

            #UNTESTED RIGHT NOW
            if len(current_workbook.sheetnames) != 3:
                logging.error('Workbook has {} tabs when 3 was expected.'.format(len(current_workbook.sheetnames)))
                move_file(x, 'error_files', cwd, 'incorrectly formatted')
                continue
            else:
                #Attempts to create file to place results log in.
                #If file exists, does not create file. Will append to file instead.
                #Opens the file and appends a separator, then closes the file again so code further down doesn't break.
                try:
                    file_name = os.path.basename(x)
                    file_name = file_name[:len(file_name) - 5] + '_result.log'
                    open(file_name, 'x')
                    logging.info('Created empty file: {}'.format(file_name))
                except:
                    logging.error('File already exists, will try to append to file instead: '.format(file_name))
                    try:
                        current_file_log = open(file_name, 'a')
                        current_file_log.write('\n\n---\n\n')
                        current_file_log.close()
                    except:
                        logging.error('Error appending to file: {}'.format(file_name))
                        continue

                selected_month = get_month(x)
                
                #Processing Sheet 1
                active_sheet = current_workbook[current_workbook.sheetnames[0]]
                row_index = 2
                for y in active_sheet.iter_rows(min_row = 2, max_row = 13, min_col = 1, max_col = 1, values_only = True):
                    #If y is date_time object
                    if type(y[0]) == datetime:
                        month = y[0].strftime('%m')
                        if int(month) == selected_month[0]+1:
                            break
                        else:
                            row_index +=1
                    #If y is str
                    elif type(y[0]) == str:
                        if y[0].lower().find(selected_month[1]) > -1:
                            break
                        else:
                            row_index +=1
                    else:
                        row_index +=1

                if row_index > 13:
                    logging.error('Error locating data on sheet {}. Skipping.'.format(current_workbook.sheetnames[0]))
                    move_file(x, 'error_files', cwd, 'incorrectly formatted')
                    current_file_log.close()
                    current_workbook.close()
                    continue

                data = []
                for y in active_sheet.iter_cols(min_row = row_index, max_row = row_index, min_col = 2, max_col = 6, values_only=True):
                    data = data + list(y)

                current_file_log = open(file_name, 'a')
                current_file_log.write('Values for {}:\n'.format(months[selected_month[0]].capitalize()))
                try:
                    current_file_log.write('\nCalls Offered : {:,}\nAbandon After 30s : {:.2%}\nFCR : {:.2%}\nDSAT : {:.2%}\nCSAT : {:.2%}\n\n'.format(data[0], data[1], data[2], data[3], data[4]))
                    logging.info('Successfully wrote sheet {} to file: {}'.format(current_workbook.sheetnames[0], file_name))
                except:
                    logging.error('Error writing sheet {} to file, list dump: {}'.format(current_workbook.sheetnames[0], data))
                    move_file(x, 'error_files', cwd, 'incorrectly formatted')
                    current_file_log.close()
                    current_workbook.close()
                    continue

                #Processing Sheet 2
                active_sheet = current_workbook[current_workbook.sheetnames[1]]
                col_index = 2

                for y in active_sheet.iter_cols(min_row = 1, max_row = 1, min_col = 2, max_col = 14, values_only = True):
                    #If y is date_time object
                    if type(y[0]) == datetime:
                        month = y[0].strftime('%m')
                        if int(month) == selected_month[0]+1:
                            break
                        else:
                            col_index +=1
                    #If y is str
                    elif type(y[0]) == str:
                        if y[0].lower().find(selected_month[1]) > -1:
                            break
                        else:
                            col_index +=1
                    else:
                        row_index +=1
                
                if col_index > 14:
                    logging.error('Error locating data on sheet {}. Skipping.'.format(current_workbook.sheetnames[1]))
                    move_file(x, 'error_files', cwd, 'incorrectly formatted')
                    current_file_log.close()
                    current_workbook.close()
                    continue

                data = []
                for y in active_sheet.iter_cols(min_row = 4, max_row = 9, min_col = col_index, max_col = col_index, values_only=True):
                    data = data + list(y)

                try:
                    current_file_log.write('Net Promoter Score:\n(Instructions file states Detractors & Passives > 100 is supposedly good for this program.)\nPromoters: {:,} - {}\nPassive: {} - {}\nDetractors: {} - {}'.format(data[0], good_bad(data[0], 200), data[2], good_bad(data[2], 100), data[4], good_bad(data[4], 100)))
                    logging.info('Successfully wrote sheet {} to file: {}'.format(current_workbook.sheetnames[1], file_name))
                    move_file(x, 'archived_files', cwd, 'successfully processed')
                    try:
                        temp = os.path.basename(x)
                        processed_file_log.write(temp + '\n')
                        logging.info('Successfully added file to processed file log: {}'.format(temp))
                    except:
                        logging.error('Could not add file to processed file log: {} '.format(temp))
                except:
                    logging.error('Error writing sheet {} to file, list dump: {}'.format(current_workbook.sheetnames[1], data))
                    move_file(x, 'error_files', cwd, 'incorrectly formatted')
                    current_file_log.close()
                    current_workbook.close()
                
                num_processed += 1
                current_file_log.close()
                current_workbook.close()
    
    processed_file_log.close()
    logging.info('{} files were successfully processed.'.format(num_processed))
    logging.info('Program exited normally.')
